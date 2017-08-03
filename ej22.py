from openpyxl import load_workbook

wb = load_workbook("datos.xlsx")

ws = wb["Hoja1"]

cells = ws["A1:C12"]

def T(cell):
    return cell.value

for row in cells:
    row_t = map(T, row)
    print row
    print row_t

    x = row_t[0]
    y = row_t[1]
    z = row_t[2]
    w = x + y + z

    print x, y, z, w

wb.close()