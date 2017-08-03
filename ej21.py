from openpyxl import load_workbook

wb = load_workbook("datos.xlsx")

ws = wb["Hoja1"]

cell = ws["B1"]

print cell.value

wb.close()