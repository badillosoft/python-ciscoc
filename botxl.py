from openpyxl import load_workbook

def extract(wb_name, ws_name, range_labels, range_data):
    wb = load_workbook(wb_name, data_only=True)
    ws = wb[ws_name]
    labels = map(lambda cell: cell.value, ws[range_labels][0])
    data = [dict(zip(labels, map(lambda cell: cell.value, row))) \
        for row in ws[range_data]]
    wb.close()
    return data

def extract_dep(wb_name, ws_name, range_labels, range_data):
    wb = load_workbook(wb_name, data_only=True)
    ws = wb[ws_name]
    def T(cell):
        return cell.value
    labels = map(T, ws[range_labels][0])
    data = []
    for row in ws[range_data]:
        row_t = map(T, row)
        D = zip(labels, row_t)
        d = {}
        for k, v in D:
            d[k] = v
        data.append(d)
    wb.close()
    return data