from openpyxl import load_workbook

wb = load_workbook("datos.xlsx")

ws = wb["Hoja2"]

# Extrar las etiquetas
def T(cell):
    return cell.value

labels = map(T, ws["B5:E5"][0])

print labels

print "-" * 40

personas = []

for row in ws["B6:E12"]:
    row_t = map(T, row)
    D = zip(labels, row_t)
    d = {}
    for k, v in D:
        d[k] = v
    print d
    personas.append(d)

print "-" * 40

print personas

wb.close()