from openpyxl import load_workbook

wb = load_workbook("datos.xlsx")

ws = wb["Hoja2"]

# Extrar las etiquetas
def T(cell):
    return cell.value

labels = map(T, ws["B5:E5"][0])

print labels

print "-" * 40

for row in ws["B6:E12"]:
    print map(T, row)

wb.close()